from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Avg
from django.db import transaction
from django.utils import timezone
from django.urls import reverse
from django.views.decorators.http import require_http_methods
import csv
import json
from io import TextIOWrapper
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import (
    Project, ProjectStatus, ProjectType, ProjectCategory, 
    ProjectPriority, Currency, ProjectHistory, ProjectAssignment
)
from authentication.decorators import permission_required
from authentication.models import UserActivity, DataFilter, Module
from authentication.utils import log_user_activity


def apply_user_data_filters(user, queryset, model_name):
    """Apply user profile data filters to a queryset"""
    if not hasattr(user, 'user_profile') or not user.user_profile.profile:
        return queryset
    
    profile = user.user_profile.profile
    
    # Get data filters for this profile and model
    try:
        module = Module.objects.get(name='project')
        filters = DataFilter.objects.filter(
            profile=profile,
            module=module,
            model_name=model_name,
            is_active=True
        )
        
        if filters.exists():
            # If there are multiple filters, we need to combine them properly
            # For data access filters, we typically want to show records that match ANY of the filters (OR logic)
            from django.db.models import Q
            combined_filter = Q()
            
            for data_filter in filters:
                if data_filter.filter_conditions:
                    try:
                        # Add each filter with OR logic
                        combined_filter |= Q(**data_filter.filter_conditions)
                    except Exception as e:
                        # Log the error but don't break the view
                        print(f"Error processing filter {data_filter.name}: {e}")
                        continue
            
            # Apply the combined filter
            if combined_filter:
                queryset = queryset.filter(combined_filter)
                    
    except Module.DoesNotExist:
        pass
    
    return queryset


@login_required
@permission_required('projects', 'view')
def project_list(request):
    """Display list of projects with filtering and pagination"""
    projects_query = Project.objects.select_related(
        'status', 'project_type', 'category', 'priority', 
        'currency', 'assigned_to', 'created_by'
    ).filter(is_active=True)
    
    # Apply user profile data filters first
    projects_query = apply_user_data_filters(request.user, projects_query, 'Project')
    
    # Apply filters
    search = request.GET.get('search')
    if search:
        projects_query = projects_query.filter(
            Q(name__icontains=search) | 
            Q(project_id__icontains=search) |
            Q(description__icontains=search) |
            Q(location__icontains=search) |
            Q(developer__icontains=search)
        )
    
    status_filter = request.GET.get('status')
    if status_filter:
        projects_query = projects_query.filter(status_id=status_filter)
    
    type_filter = request.GET.get('type')
    if type_filter:
        projects_query = projects_query.filter(project_type_id=type_filter)
    
    category_filter = request.GET.get('category')
    if category_filter:
        projects_query = projects_query.filter(category_id=category_filter)
    
    priority_filter = request.GET.get('priority')
    if priority_filter:
        projects_query = projects_query.filter(priority_id=priority_filter)
    
    assigned_filter = request.GET.get('assigned')
    if assigned_filter:
        if assigned_filter == 'me':
            projects_query = projects_query.filter(assigned_to=request.user)
        elif assigned_filter == 'unassigned':
            projects_query = projects_query.filter(assigned_to__isnull=True)
        else:
            projects_query = projects_query.filter(assigned_to_id=assigned_filter)
    
    # Sorting
    sort_by = request.GET.get('sort', 'created_at')
    sort_order = request.GET.get('order', 'desc')
    if sort_order == 'desc':
        sort_by = f'-{sort_by}'
    
    projects_query = projects_query.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(projects_query, 20)
    page_number = request.GET.get('page')
    projects = paginator.get_page(page_number)
    
    # Get filter options
    statuses = ProjectStatus.objects.filter(is_active=True).order_by('order')
    types = ProjectType.objects.filter(is_active=True).order_by('order')
    categories = ProjectCategory.objects.filter(is_active=True).order_by('order')
    priorities = ProjectPriority.objects.filter(is_active=True).order_by('order')
    
    # Statistics
    stats = {
        'total_projects': Project.objects.filter(is_active=True).count(),
        'active_projects': Project.objects.filter(is_active=True, status__name='active').count(),
        'total_units': Project.objects.filter(is_active=True).aggregate(
            total=Sum('total_units'))['total'] or 0,
        'available_units': Project.objects.filter(is_active=True).aggregate(
            available=Sum('available_units'))['available'] or 0,
    }
    
    # Log activity
    log_user_activity(request.user, 'view', 'projects', 'Viewed project list')
    
    context = {
        'projects': projects,
        'statuses': statuses,
        'types': types,
        'categories': categories,
        'priorities': priorities,
        'stats': stats,
        'search': search,
        'current_filters': {
            'status': status_filter,
            'type': type_filter,
            'category': category_filter,
            'priority': priority_filter,
            'assigned': assigned_filter,
        },
        'sort_by': request.GET.get('sort', 'created_at'),
        'sort_order': sort_order,
    }
    
    return render(request, 'projects/project_list.html', context)


@login_required
@permission_required('projects', 'view')
def project_detail(request, project_id):
    """Display detailed view of a project"""
    project = get_object_or_404(
        Project.objects.select_related(
            'status', 'project_type', 'category', 'priority', 
            'currency', 'assigned_to', 'created_by'
        ),
        project_id=project_id,
        is_active=True
    )
    
    # Get project history
    history = ProjectHistory.objects.filter(project=project).order_by('-created_at')[:20]
    
    # Get project assignments
    assignments = ProjectAssignment.objects.filter(
        project=project, 
        is_active=True
    ).select_related('user', 'assigned_by')
    
    # Log activity
    log_user_activity(request.user, 'view', 'projects', f'Viewed project {project.name}')
    
    context = {
        'project': project,
        'history': history,
        'assignments': assignments,
    }
    
    return render(request, 'projects/project_detail.html', context)


@login_required
@permission_required('projects', 'create')
def project_create(request):
    """Create a new project"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Create project
                project = Project.objects.create(
                    name=request.POST.get('name'),
                    description=request.POST.get('description', ''),
                    location=request.POST.get('location', ''),
                    developer=request.POST.get('developer', ''),
                    status_id=request.POST.get('status'),
                    project_type_id=request.POST.get('project_type'),
                    category_id=request.POST.get('category') if request.POST.get('category') else None,
                    priority_id=request.POST.get('priority') if request.POST.get('priority') else None,
                    start_date=request.POST.get('start_date') if request.POST.get('start_date') else None,
                    end_date=request.POST.get('end_date') if request.POST.get('end_date') else None,
                    completion_year=request.POST.get('completion_year') if request.POST.get('completion_year') else None,
                    total_units=request.POST.get('total_units') if request.POST.get('total_units') else None,
                    available_units=request.POST.get('available_units') if request.POST.get('available_units') else None,
                    price_range=request.POST.get('price_range', ''),
                    currency_id=request.POST.get('currency') if request.POST.get('currency') else None,
                    min_price=request.POST.get('min_price') if request.POST.get('min_price') else None,
                    max_price=request.POST.get('max_price') if request.POST.get('max_price') else None,
                    assigned_to_id=request.POST.get('assigned_to') if request.POST.get('assigned_to') else None,
                    notes=request.POST.get('notes', ''),
                    tags=request.POST.get('tags', ''),
                    is_featured=request.POST.get('is_featured') == 'on',
                    created_by=request.user
                )
                
                # Log activity
                log_user_activity(request.user, 'create', 'projects', f'Created project {project.name}')
                
                # Create history entry
                ProjectHistory.objects.create(
                    project=project,
                    action='created',
                    field_name='project',
                    new_value='Project created',
                    user=request.user
                )
                
                messages.success(request, f'Project "{project.name}" was created successfully!')
                return redirect('projects:project_detail', project_id=project.project_id)
                
        except Exception as e:
            messages.error(request, f'Error creating project: {str(e)}')
    
    # Get form options
    statuses = ProjectStatus.objects.filter(is_active=True).order_by('order')
    types = ProjectType.objects.filter(is_active=True).order_by('order')
    categories = ProjectCategory.objects.filter(is_active=True).order_by('order')
    priorities = ProjectPriority.objects.filter(is_active=True).order_by('order')
    currencies = Currency.objects.filter(is_active=True).order_by('order')
    
    context = {
        'statuses': statuses,
        'types': types,
        'categories': categories,
        'priorities': priorities,
        'currencies': currencies,
    }
    
    return render(request, 'projects/project_create.html', context)


@login_required
@permission_required('projects', 'edit')
def project_edit(request, project_id):
    """Edit an existing project"""
    project = get_object_or_404(Project, project_id=project_id, is_active=True)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Track changes for history
                changes = []
                
                fields_to_track = {
                    'name': 'Name',
                    'description': 'Description',
                    'location': 'Location',
                    'developer': 'Developer',
                    'status_id': 'Status',
                    'project_type_id': 'Project Type',
                    'category_id': 'Category',
                    'priority_id': 'Priority',
                    'start_date': 'Start Date',
                    'end_date': 'End Date',
                    'completion_year': 'Completion Year',
                    'total_units': 'Total Units',
                    'available_units': 'Available Units',
                    'price_range': 'Price Range',
                    'currency_id': 'Currency',
                    'min_price': 'Min Price',
                    'max_price': 'Max Price',
                    'assigned_to_id': 'Assigned To',
                    'notes': 'Notes',
                    'tags': 'Tags',
                    'is_featured': 'Is Featured',
                }
                
                for field, display_name in fields_to_track.items():
                    old_value = getattr(project, field)
                    new_value = request.POST.get(field.replace('_id', ''))
                    
                    if field == 'is_featured':
                        new_value = request.POST.get('is_featured') == 'on'
                    elif field.endswith('_id') and new_value:
                        new_value = int(new_value) if new_value else None
                    elif field in ['total_units', 'available_units', 'completion_year'] and new_value:
                        new_value = int(new_value) if new_value else None
                    elif field in ['min_price', 'max_price'] and new_value:
                        new_value = float(new_value) if new_value else None
                    elif new_value == '':
                        new_value = None
                    
                    if old_value != new_value:
                        changes.append({
                            'field': field,
                            'display_name': display_name,
                            'old_value': old_value,
                            'new_value': new_value
                        })
                
                # Update project
                project.name = request.POST.get('name')
                project.description = request.POST.get('description', '')
                project.location = request.POST.get('location', '')
                project.developer = request.POST.get('developer', '')
                project.status_id = request.POST.get('status')
                project.project_type_id = request.POST.get('project_type')
                project.category_id = request.POST.get('category') if request.POST.get('category') else None
                project.priority_id = request.POST.get('priority') if request.POST.get('priority') else None
                project.start_date = request.POST.get('start_date') if request.POST.get('start_date') else None
                project.end_date = request.POST.get('end_date') if request.POST.get('end_date') else None
                project.completion_year = request.POST.get('completion_year') if request.POST.get('completion_year') else None
                project.total_units = request.POST.get('total_units') if request.POST.get('total_units') else None
                project.available_units = request.POST.get('available_units') if request.POST.get('available_units') else None
                project.price_range = request.POST.get('price_range', '')
                project.currency_id = request.POST.get('currency') if request.POST.get('currency') else None
                project.min_price = request.POST.get('min_price') if request.POST.get('min_price') else None
                project.max_price = request.POST.get('max_price') if request.POST.get('max_price') else None
                project.assigned_to_id = request.POST.get('assigned_to') if request.POST.get('assigned_to') else None
                project.notes = request.POST.get('notes', '')
                project.tags = request.POST.get('tags', '')
                project.is_featured = request.POST.get('is_featured') == 'on'
                
                project.save()
                
                # Log changes
                for change in changes:
                    ProjectHistory.objects.create(
                        project=project,
                        action='updated',
                        field_name=change['field'],
                        old_value=str(change['old_value']) if change['old_value'] is not None else '',
                        new_value=str(change['new_value']) if change['new_value'] is not None else '',
                        user=request.user
                    )
                
                # Log activity
                log_user_activity(request.user, 'edit', 'projects', f'Updated project {project.name}')
                
                messages.success(request, f'Project "{project.name}" was updated successfully!')
                return redirect('projects:project_detail', project_id=project.project_id)
                
        except Exception as e:
            messages.error(request, f'Error updating project: {str(e)}')
    
    # Get form options
    statuses = ProjectStatus.objects.filter(is_active=True).order_by('order')
    types = ProjectType.objects.filter(is_active=True).order_by('order')
    categories = ProjectCategory.objects.filter(is_active=True).order_by('order')
    priorities = ProjectPriority.objects.filter(is_active=True).order_by('order')
    currencies = Currency.objects.filter(is_active=True).order_by('order')
    
    context = {
        'project': project,
        'statuses': statuses,
        'types': types,
        'categories': categories,
        'priorities': priorities,
        'currencies': currencies,
    }
    
    return render(request, 'projects/project_edit.html', context)


@login_required
@permission_required('projects', 'delete')
@require_http_methods(["POST"])
def project_delete(request, project_id):
    """Soft delete a project"""
    project = get_object_or_404(Project, project_id=project_id, is_active=True)
    
    try:
        project.is_active = False
        project.save()
        
        # Log activity
        log_user_activity(request.user, 'delete', 'projects', f'Deleted project {project.name}')
        
        # Create history entry
        ProjectHistory.objects.create(
            project=project,
            action='deleted',
            field_name='is_active',
            old_value='True',
            new_value='False',
            user=request.user
        )
        
        messages.success(request, f'Project "{project.name}" was deleted successfully!')
        
    except Exception as e:
        messages.error(request, f'Error deleting project: {str(e)}')
    
    return redirect('projects:project_list')


@login_required
@permission_required('projects', 'export')
def project_export(request):
    """Export projects to Excel file"""
    try:
        # Get filtered projects
        projects_query = Project.objects.select_related(
            'status', 'project_type', 'category', 'priority', 
            'currency', 'assigned_to', 'created_by'
        ).filter(is_active=True)
        
        # Apply same filters as list view
        search = request.GET.get('search')
        if search:
            projects_query = projects_query.filter(
                Q(name__icontains=search) | 
                Q(project_id__icontains=search) |
                Q(description__icontains=search) |
                Q(location__icontains=search) |
                Q(developer__icontains=search)
            )
        
        status_filter = request.GET.get('status')
        if status_filter:
            projects_query = projects_query.filter(status_id=status_filter)
        
        type_filter = request.GET.get('type')
        if type_filter:
            projects_query = projects_query.filter(project_type_id=type_filter)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Projects"
        
        # Headers
        headers = [
            'Project ID', 'Name', 'Description', 'Location', 'Developer',
            'Status', 'Type', 'Category', 'Priority', 'Start Date', 'End Date',
            'Completion Year', 'Total Units', 'Available Units', 'Price Range',
            'Currency', 'Min Price', 'Max Price', 'Assigned To', 'Created By',
            'Created At', 'Notes', 'Tags'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, project in enumerate(projects_query, 2):
            data = [
                project.project_id,
                project.name,
                project.description,
                project.location,
                project.developer,
                project.status.display_name if project.status else '',
                project.project_type.display_name if project.project_type else '',
                project.category.display_name if project.category else '',
                project.priority.display_name if project.priority else '',
                project.start_date.strftime('%Y-%m-%d') if project.start_date else '',
                project.end_date.strftime('%Y-%m-%d') if project.end_date else '',
                project.completion_year,
                project.total_units,
                project.available_units,
                project.price_range,
                project.currency.code if project.currency else '',
                project.min_price,
                project.max_price,
                project.assigned_to.get_full_name() if project.assigned_to else '',
                project.created_by.get_full_name() if project.created_by else '',
                project.created_at.strftime('%Y-%m-%d %H:%M'),
                project.notes,
                project.tags
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="projects_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        
        # Log activity
        log_user_activity(request.user, 'export', 'projects', f'Exported {projects_query.count()} projects')
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error exporting projects: {str(e)}')
        return redirect('projects:project_list')


@login_required
@permission_required('projects', 'import')
def project_import(request):
    """Import projects from CSV/Excel file"""
    if request.method == 'POST':
        try:
            file = request.FILES.get('file')
            if not file:
                messages.error(request, 'Please select a file to import.')
                return redirect('projects:project_list')
            
            # Validate file type
            if not (file.name.endswith('.csv') or file.name.endswith('.xlsx')):
                messages.error(request, 'Please upload a CSV or Excel file.')
                return redirect('projects:project_list')
            
            imported_count = 0
            error_count = 0
            errors = []
            
            with transaction.atomic():
                if file.name.endswith('.csv'):
                    # Handle CSV file
                    file_wrapper = TextIOWrapper(file.file, encoding='utf-8')
                    csv_reader = csv.DictReader(file_wrapper)
                    
                    for row_num, row in enumerate(csv_reader, 2):
                        try:
                            # Create project from CSV row
                            project = Project.objects.create(
                                name=row.get('Name', '').strip(),
                                description=row.get('Description', '').strip(),
                                location=row.get('Location', '').strip(),
                                developer=row.get('Developer', '').strip(),
                                price_range=row.get('Price Range', '').strip(),
                                notes=row.get('Notes', '').strip(),
                                tags=row.get('Tags', '').strip(),
                                created_by=request.user
                            )
                            
                            # Log history
                            ProjectHistory.objects.create(
                                project=project,
                                action='imported',
                                field_name='project',
                                new_value='Project imported from CSV',
                                user=request.user
                            )
                            
                            imported_count += 1
                            
                        except Exception as e:
                            error_count += 1
                            errors.append(f'Row {row_num}: {str(e)}')
                            if error_count >= 10:  # Limit error messages
                                errors.append('... and more errors')
                                break
                
                # Log activity
                log_user_activity(request.user, 'import', 'projects', f'Imported {imported_count} projects')
                
                if imported_count > 0:
                    messages.success(request, f'Successfully imported {imported_count} projects!')
                
                if error_count > 0:
                    error_msg = f'{error_count} errors occurred during import.'
                    if errors:
                        error_msg += ' First few errors: ' + '; '.join(errors[:5])
                    messages.warning(request, error_msg)
                
        except Exception as e:
            messages.error(request, f'Error importing projects: {str(e)}')
    
    return redirect('projects:project_list')
